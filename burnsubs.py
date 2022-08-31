import pysubs2
from pysubs2 import SSAFile
import ffmpeg
from openpyxl import Workbook,load_workbook

class srt2ass():
    def __init__(self,source,font,fontsize,marginv,resX,resY,bold):
        self.source = source
        self.font = font
        self.fontsize = fontsize
        self.marginv = marginv
        self.resX = resX
        self.resY = resY
        self.bold = bold
    def transfer(self):
        self.newsubs = SSAFile.load(self.source)
        self.cur_style = pysubs2.SSAStyle(self.font,self.fontsize,bold=self.bold)
        self.cur_style.marginv = self.marginv
        self.cur_style.marginl = self.cur_style.marginr = 40
        self.newsubs.styles["Default"] = self.cur_style
        self.newsubs.save("trans-"+ source.split('.')[0] + ".ass")

        self.assname = "trans-"+ source.split('.')[0] + ".ass"

        self.add_res()

    def add_res(self):
        res_info = "PlayResX: " + self.resX + '\n' + "PlayResY: " +self.resY +'\n'
        fp = open(self.assname, encoding="utf-8")
        lines = []
        for line in fp:
            lines.append(line)
        fp.close()
        lines.insert(6, res_info)  # 在第二行插入
        s = ''.join(lines)
        fp = open(self.assname, 'w', encoding="utf-8")
        fp.write(s)
        fp.close()

class burnsubs(srt2ass):
    def __init__(self,source, font, fontsize, marginv, resX, resY, in_video, out_video,crf,bold):
        srt2ass.__init__(self, source, font, fontsize, marginv, resX, resY,bold)
        self.in_video = in_video
        self.out_video = out_video
        self.crf = crf

    def burn(self):
        self.transfer()
        video = ffmpeg.input(self.in_video)
        audio = video.audio
        ffmpeg.concat(video.filter("subtitles", self.assname), audio, v=1, a=1)\
            .output(self.out_video, crf = self.crf).run()
# resX = '1920'
# resY = '1080'
# source = 'id.srt'
# font = '微软雅黑'
# fontsize = 54
# # marginl = marginr = 135
# marginv = 45
# in_video = 'id.mp4'
# out_video = 'addsub_id.mp4'
# crf = 16y
wb = load_workbook('ass_styles.xlsx')
ws = wb['Sheet1']
langparas = []
# print(ws.max_row)
for i in range (2,ws.max_row+1):
    lang = []
    for j in range(2,ws.max_column+1):
        lang.append(ws.cell(i,j).value)
    langparas.append(lang)
print(langparas)

for i in range(0,len(langparas)):
    source =langparas[i][0]#字幕
    font = langparas[i][1]#字体
    fontsize = langparas[i][2]#字号
    marginv = langparas[i][3]#垂直边距
    resX = str(langparas[i][4])#水平分辨率
    resY = str(langparas[i][5])#垂直分辨率
    in_video = langparas[i][6]#原视频
    out_video = langparas[i][7]#生成视频
    crf = langparas[i][8]
    bold = langparas[i][9]


    print(source,font,fontsize,marginv,resX,resY,in_video,out_video,crf,bold)
    a = burnsubs(source,font,fontsize,marginv,resX,resY,in_video,out_video,crf,bold)
    a.burn()


input("压制完成")
